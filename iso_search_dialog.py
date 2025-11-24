# iso_search_dialog.py

import os
import csv
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem,
    QHeaderView, QDialogButtonBox, QPushButton, QLineEdit, QFileDialog,
    QMessageBox, QMenu, QProgressDialog, QApplication, QGroupBox, QCheckBox, QWidget
)
from PyQt6.QtCore import Qt, QSize, pyqtSignal, QTimer
from PyQt6.QtGui import QIcon, QKeySequence, QShortcut, QAction


class IsoSearchDialog(QDialog):
    """
    دیالوگ پیشرفته جستجو و مدیریت فایل‌های ISO/DWG

    ویژگی‌ها:
    - جستجوی متنی real-time
    - مرتب‌سازی ستون‌ها
    - انتخاب چندتایی و باز کردن دسته‌ای
    - کپی مسیر و باز کردن پوشه
    - نمایش اطلاعات کامل (سایز، تاریخ، نوع)
    - Export به CSV/Excel
    - میانبرهای صفحه‌کلید
    """

    # سیگنال برای اطلاع‌رسانی به پنجره اصلی
    files_opened = pyqtSignal(list)  # لیست فایل‌های باز شده

    def __init__(self, data_manager, line_no: str, parent=None):
        super().__init__(parent)
        self.dm = data_manager
        self.line_no = line_no
        self.parent_window = parent
        self.matches: List[str] = []
        self.filtered_matches: List[str] = []
        self.file_info_cache: Dict[str, Dict] = {}  # کش اطلاعات فایل‌ها

        # تنظیمات دیالوگ
        self.setWindowTitle(f"جستجوی فایل‌های ISO/DWG - Line: {line_no}")
        self.resize(1200, 650)
        self.setMinimumSize(900, 500)

        self.setup_ui()
        self.setup_shortcuts()
        self.perform_search()
        self.apply_styles()

    def setup_ui(self):
        """ساخت رابط کاربری کامل دیالوگ"""
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(10)

        # === بخش بالا: آمار و جستجو ===
        top_section = self._create_top_section()
        main_layout.addWidget(top_section)

        # === بخش میانی: جدول نتایج ===
        self.table = self._create_results_table()
        main_layout.addWidget(self.table)

        # === بخش پایین: دکمه‌های عملیاتی ===
        bottom_section = self._create_bottom_section()
        main_layout.addWidget(bottom_section)

    def _create_top_section(self) -> QGroupBox:
        """ساخت بخش بالایی شامل آمار و جستجو"""
        group = QGroupBox("اطلاعات و فیلتر")
        layout = QVBoxLayout(group)

        # ردیف اول: آمار
        stats_layout = QHBoxLayout()
        self.stats_label = QLabel("در حال جستجو...")
        self.stats_label.setStyleSheet("font-weight: bold; color: #2196F3;")
        stats_layout.addWidget(self.stats_label)
        stats_layout.addStretch()

        # دکمه Refresh
        self.refresh_btn = QPushButton("🔄 بازخوانی")
        self.refresh_btn.setToolTip("جستجوی مجدد فایل‌ها (F5)")
        self.refresh_btn.clicked.connect(self.perform_search)
        stats_layout.addWidget(self.refresh_btn)

        layout.addLayout(stats_layout)

        # ردیف دوم: جستجوی متنی
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 فیلتر:"))

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("نام فایل یا مسیر را وارد کنید...")
        self.search_input.setClearButtonEnabled(True)
        self.search_input.textChanged.connect(self._on_search_text_changed)
        search_layout.addWidget(self.search_input, 1)

        # دکمه Clear Filter
        self.clear_filter_btn = QPushButton("❌ پاک کردن")
        self.clear_filter_btn.clicked.connect(lambda: self.search_input.clear())
        search_layout.addWidget(self.clear_filter_btn)

        layout.addLayout(search_layout)

        # چک‌باکس نمایش پوشه‌ها
        self.show_folders_cb = QCheckBox("نمایش ستون مسیر کامل")
        self.show_folders_cb.setChecked(True)
        self.show_folders_cb.stateChanged.connect(self._toggle_folder_column)
        layout.addWidget(self.show_folders_cb)

        return group

    def _create_results_table(self) -> QTableWidget:
        """ساخت جدول نتایج با قابلیت‌های پیشرفته"""
        table = QTableWidget(0, 6)

        # تنظیم هدرها
        headers = ["نام فایل", "نوع", "حجم", "تاریخ تغییر", "مسیر پوشه", "مسیر کامل"]
        table.setHorizontalHeaderLabels(headers)

        # تنظیمات ظاهری
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)  # انتخاب چندتایی
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSortingEnabled(True)  # فعال‌سازی مرتب‌سازی

        # تنظیم عرض ستون‌ها
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)  # نام
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # نوع
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # حجم
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # تاریخ
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # مسیر پوشه
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Interactive)  # مسیر کامل

        # مخفی کردن ستون مسیر کامل (فقط برای ذخیره داده)
        table.setColumnHidden(5, True)

        # اتصال رویدادها
        table.cellDoubleClicked.connect(self._on_double_click)
        table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        table.customContextMenuRequested.connect(self._show_context_menu)

        return table

    def _create_bottom_section(self) -> QWidget:
        """ساخت بخش پایینی شامل دکمه‌های عملیاتی"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)

        # ردیف اول: دکمه‌های اصلی
        main_buttons_layout = QHBoxLayout()

        self.open_btn = QPushButton("📂 باز کردن")
        self.open_btn.setToolTip("باز کردن فایل(های) انتخابی (Enter)")
        self.open_btn.clicked.connect(self.open_selected_files)
        self.open_btn.setEnabled(False)

        self.open_folder_btn = QPushButton("📁 باز کردن پوشه")
        self.open_folder_btn.setToolTip("باز کردن پوشه حاوی فایل (Ctrl+E)")
        self.open_folder_btn.clicked.connect(self.open_containing_folder)
        self.open_folder_btn.setEnabled(False)

        self.copy_path_btn = QPushButton("📋 کپی مسیر")
        self.copy_path_btn.setToolTip("کپی مسیر فایل(ها) (Ctrl+C)")
        self.copy_path_btn.clicked.connect(self.copy_selected_paths)
        self.copy_path_btn.setEnabled(False)

        self.select_all_btn = QPushButton("☑️ انتخاب همه")
        self.select_all_btn.setToolTip("انتخاب تمام فایل‌های نمایش داده شده (Ctrl+A)")
        self.select_all_btn.clicked.connect(self.table.selectAll)

        main_buttons_layout.addWidget(self.open_btn)
        main_buttons_layout.addWidget(self.open_folder_btn)
        main_buttons_layout.addWidget(self.copy_path_btn)
        main_buttons_layout.addWidget(self.select_all_btn)
        main_buttons_layout.addStretch()

        layout.addLayout(main_buttons_layout)

        # ردیف دوم: Export و Close
        bottom_buttons_layout = QHBoxLayout()

        self.export_btn = QPushButton("💾 Export به CSV")
        self.export_btn.setToolTip("ذخیره لیست فایل‌ها در فایل CSV")
        self.export_btn.clicked.connect(self.export_to_csv)

        self.export_excel_btn = QPushButton("📊 Export به Excel")
        self.export_excel_btn.setToolTip("ذخیره لیست فایل‌ها در فایل Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel)

        bottom_buttons_layout.addWidget(self.export_btn)
        bottom_buttons_layout.addWidget(self.export_excel_btn)
        bottom_buttons_layout.addStretch()

        self.close_btn = QPushButton("❌ بستن")
        self.close_btn.clicked.connect(self.reject)
        bottom_buttons_layout.addWidget(self.close_btn)

        layout.addLayout(bottom_buttons_layout)

        # لیبل انتخاب شده‌ها
        self.selection_label = QLabel("هیچ فایلی انتخاب نشده")
        self.selection_label.setStyleSheet("color: #666; font-size: 11px;")
        layout.addWidget(self.selection_label)

        # اتصال سیگنال تغییر انتخاب
        self.table.itemSelectionChanged.connect(self._on_selection_changed)

        return widget

    def setup_shortcuts(self):
        """تعریف میانبرهای صفحه‌کلید"""
        # Enter/Return - باز کردن فایل
        QShortcut(QKeySequence(Qt.Key.Key_Return), self, self.open_selected_files)
        QShortcut(QKeySequence(Qt.Key.Key_Enter), self, self.open_selected_files)

        # Ctrl+C - کپی مسیر
        QShortcut(QKeySequence.StandardKey.Copy, self, self.copy_selected_paths)

        # Ctrl+A - انتخاب همه
        QShortcut(QKeySequence.StandardKey.SelectAll, self, self.table.selectAll)

        # Ctrl+E - باز کردن پوشه
        QShortcut(QKeySequence("Ctrl+E"), self, self.open_containing_folder)

        # F5 - رفرش
        QShortcut(QKeySequence(Qt.Key.Key_F5), self, self.perform_search)

        # Ctrl+F - فوکوس روی جستجو
        QShortcut(QKeySequence.StandardKey.Find, self, lambda: self.search_input.setFocus())

        # Escape - بستن دیالوگ
        QShortcut(QKeySequence(Qt.Key.Key_Escape), self, self.reject)

    def apply_styles(self):
        """اعمال استایل‌های CSS به دیالوگ"""
        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ccc;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QPushButton {
                padding: 6px 12px;
                border-radius: 4px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #e3f2fd;
            }
            QPushButton:disabled {
                color: #999;
            }
            QLineEdit {
                padding: 6px;
                border: 1px solid #ccc;
                border-radius: 4px;
            }
            QLineEdit:focus {
                border: 1px solid #2196F3;
            }
            QTableWidget {
                gridline-color: #e0e0e0;
                selection-background-color: #2196F3;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 6px;
                border: none;
                border-right: 1px solid #ddd;
                border-bottom: 1px solid #ddd;
                font-weight: bold;
            }
        """)

    # ===== عملیات جستجو و پردازش =====

    def perform_search(self):
        """جستجوی فایل‌های ISO و پر کردن جدول"""
        self.table.setSortingEnabled(False)  # غیرفعال کردن موقت برای سرعت
        self.table.setRowCount(0)
        self.file_info_cache.clear()

        # نمایش وضعیت در حال بارگذاری
        self.stats_label.setText("🔄 در حال جستجو...")
        QApplication.processEvents()

        try:
            # جستجو در دیتابیس
            self.matches = self.dm.find_iso_files(self.line_no)
            self.filtered_matches = self.matches.copy()

        except Exception as e:
            self._log_to_parent(f"❌ جستجوی فایل‌ها با خطا مواجه شد: {e}", "error")
            self.matches = []
            self.filtered_matches = []
            self.stats_label.setText("❌ خطا در جستجو")
            return

        if not self.matches:
            self._log_to_parent("⚠️ فایلی مطابق با Line No پیدا نشد.", "warning")
            self.stats_label.setText("⚠️ فایلی یافت نشد")
            self._show_no_results_message()
            return

        # بارگذاری داده‌ها
        self._populate_table(self.filtered_matches)
        self._update_stats()
        self._log_to_parent(f"✅ {len(self.matches)} فایل پیدا شد.", "success")

        self.table.setSortingEnabled(True)  # فعال‌سازی مجدد مرتب‌سازی

    def _populate_table(self, file_paths: List[str]):
        """پر کردن جدول با اطلاعات فایل‌ها"""
        self.table.setRowCount(len(file_paths))

        # استفاده از progress dialog برای فایل‌های زیاد
        show_progress = len(file_paths) > 50
        progress = None

        if show_progress:
            progress = QProgressDialog("در حال بارگذاری اطلاعات...", "لغو", 0, len(file_paths), self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)

        for row, file_path in enumerate(file_paths):
            if show_progress:
                progress.setValue(row)
                if progress.wasCanceled():
                    break
                QApplication.processEvents()

            # دریافت اطلاعات فایل
            info = self._get_file_info(file_path)

            # ستون 0: نام فایل
            name_item = QTableWidgetItem(info['name'])
            name_item.setToolTip(info['name'])
            self.table.setItem(row, 0, name_item)

            # ستون 1: نوع فایل
            type_item = QTableWidgetItem(info['type'])
            type_item.setToolTip(f"نوع: {info['type']}")
            self.table.setItem(row, 1, type_item)

            # ستون 2: حجم
            size_item = QTableWidgetItem(info['size_str'])
            size_item.setData(Qt.ItemDataRole.UserRole, info['size_bytes'])  # برای مرتب‌سازی
            size_item.setToolTip(f"حجم: {info['size_str']}")
            self.table.setItem(row, 2, size_item)

            # ستون 3: تاریخ تغییر
            date_item = QTableWidgetItem(info['modified_str'])
            date_item.setData(Qt.ItemDataRole.UserRole, info['modified_timestamp'])
            date_item.setToolTip(f"آخرین تغییر: {info['modified_str']}")
            self.table.setItem(row, 3, date_item)

            # ستون 4: مسیر پوشه
            folder_item = QTableWidgetItem(info['folder'])
            folder_item.setToolTip(info['folder'])
            self.table.setItem(row, 4, folder_item)

            # ستون 5: مسیر کامل (مخفی)
            full_path_item = QTableWidgetItem(file_path)
            self.table.setItem(row, 5, full_path_item)

        if show_progress:
            progress.setValue(len(file_paths))

    def _get_file_info(self, file_path: str) -> Dict:
        """دریافت اطلاعات کامل یک فایل با کش"""
        if file_path in self.file_info_cache:
            return self.file_info_cache[file_path]

        path_obj = Path(file_path)

        try:
            stat = path_obj.stat()
            size_bytes = stat.st_size
            modified_timestamp = stat.st_mtime

            # فرمت‌بندی حجم
            size_str = self._format_file_size(size_bytes)

            # فرمت‌بندی تاریخ
            modified_dt = datetime.fromtimestamp(modified_timestamp)
            modified_str = modified_dt.strftime('%Y/%m/%d %H:%M')

        except (OSError, FileNotFoundError):
            size_bytes = 0
            size_str = "نامشخص"
            modified_timestamp = 0
            modified_str = "نامشخص"

        info = {
            'name': path_obj.name,
            'type': path_obj.suffix.upper().replace('.', '') or 'فایل',
            'size_bytes': size_bytes,
            'size_str': size_str,
            'modified_timestamp': modified_timestamp,
            'modified_str': modified_str,
            'folder': str(path_obj.parent),
            'full_path': file_path
        }

        self.file_info_cache[file_path] = info
        return info

    @staticmethod
    def _format_file_size(size_bytes: int) -> str:
        """فرمت‌بندی حجم فایل به واحد مناسب"""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024.0:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024.0
        return f"{size_bytes:.1f} TB"

    def _show_no_results_message(self):
        """نمایش پیام عدم یافتن نتیجه"""
        self.table.setRowCount(1)
        no_result_item = QTableWidgetItem("⚠️ هیچ فایلی با این مشخصات پیدا نشد")
        no_result_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
        no_result_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setItem(0, 0, no_result_item)
        self.table.setSpan(0, 0, 1, 6)

    # ===== فیلتر و جستجو =====

    def _on_search_text_changed(self, text: str):
        """فیلتر کردن نتایج بر اساس متن جستجو"""
        if not text.strip():
            self.filtered_matches = self.matches.copy()
        else:
            search_lower = text.lower()
            self.filtered_matches = [
                path for path in self.matches
                if search_lower in path.lower()
            ]

        self.table.setSortingEnabled(False)
        self._populate_table(self.filtered_matches)
        self._update_stats()
        self.table.setSortingEnabled(True)

    def _toggle_folder_column(self, state):
        """نمایش/مخفی کردن ستون مسیر پوشه"""
        self.table.setColumnHidden(4, not state)

    def _update_stats(self):
        """بروزرسانی لیبل آمار"""
        total = len(self.matches)
        filtered = len(self.filtered_matches)

        if filtered < total:
            self.stats_label.setText(
                f"📊 {filtered} از {total} فایل نمایش داده شده"
            )
        else:
            self.stats_label.setText(f"📊 {total} فایل یافت شد")

    # ===== عملیات روی فایل‌ها =====

    def _get_selected_file_paths(self) -> List[str]:
        """دریافت لیست مسیرهای فایل‌های انتخاب شده"""
        selected_rows = set(item.row() for item in self.table.selectedItems())
        paths = []

        for row in selected_rows:
            path_item = self.table.item(row, 5)  # ستون مسیر کامل
            if path_item:
                paths.append(path_item.text())

        return paths

    def open_selected_files(self):
        """باز کردن فایل(های) انتخاب شده"""
        paths = self._get_selected_file_paths()

        if not paths:
            return

        # هشدار برای باز کردن تعداد زیاد فایل
        if len(paths) > 10:
            reply = QMessageBox.question(
                self,
                "تایید عملیات",
                f"شما در حال باز کردن {len(paths)} فایل هستید.\n"
                "این ممکن است زمان‌بر باشد. ادامه می‌دهید؟",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # باز کردن فایل‌ها
        success_count = 0
        failed_files = []

        for path in paths:
            try:
                os.startfile(path)
                success_count += 1
                self._log_to_parent(f"📂 فایل باز شد: {Path(path).name}", "info")
            except Exception as e:
                failed_files.append((path, str(e)))
                self._log_to_parent(f"❌ خطا در باز کردن {Path(path).name}: {e}", "error")

        # نمایش نتیجه
        if failed_files:
            error_msg = "\n".join([f"• {Path(p).name}: {e}" for p, e in failed_files])
            QMessageBox.warning(
                self,
                "خطا در باز کردن فایل‌ها",
                f"{success_count} فایل با موفقیت باز شد.\n\n"
                f"خطا در {len(failed_files)} فایل:\n{error_msg}"
            )
        elif success_count > 1:
            self._log_to_parent(f"✅ {success_count} فایل با موفقیت باز شد", "success")

        # ارسال سیگنال
        self.files_opened.emit(paths)

    def open_containing_folder(self):
        """باز کردن پوشه حاوی فایل انتخاب شده"""
        paths = self._get_selected_file_paths()

        if not paths:
            return

        # باز کردن اولین پوشه
        first_folder = str(Path(paths[0]).parent)
        try:
            os.startfile(first_folder)
            self._log_to_parent(f"📁 پوشه باز شد: {first_folder}", "info")
        except Exception as e:
            self._log_to_parent(f"❌ خطا در باز کردن پوشه: {e}", "error")
            QMessageBox.critical(self, "خطا", f"خطا در باز کردن پوشه:\n{e}")

    def copy_selected_paths(self):
        """کپی کردن مسیر فایل(های) انتخاب شده به کلیپبورد"""
        paths = self._get_selected_file_paths()

        if not paths:
            return

        # کپی به کلیپبورد
        clipboard_text = "\n".join(paths)
        clipboard = QApplication.clipboard()
        clipboard.setText(clipboard_text)

        # پیام تایید
        count = len(paths)
        message = f"مسیر {count} فایل کپی شد" if count > 1 else "مسیر فایل کپی شد"
        self._log_to_parent(f"📋 {message}", "success")

        # نمایش موقت در status
        original_text = self.selection_label.text()
        self.selection_label.setText(f"✅ {message}")
        self.selection_label.setStyleSheet("color: #4CAF50; font-weight: bold;")

        QTimer.singleShot(2000, lambda: (
            self.selection_label.setText(original_text),
            self.selection_label.setStyleSheet("color: #666; font-size: 11px;")
        ))

    # ===== Export =====

    def export_to_csv(self):
        """Export لیست فایل‌ها به فرمت CSV"""
        if not self.filtered_matches:
            QMessageBox.information(self, "اطلاع", "فایلی برای Export وجود ندارد.")
            return

        # انتخاب مسیر ذخیره
        default_name = f"ISO_Files_{self.line_no.replace('/', '-')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "ذخیره به CSV",
            default_name,
            "CSV Files (*.csv)"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)

                # نوشتن هدر
                writer.writerow(['نام فایل', 'نوع', 'حجم', 'تاریخ تغییر', 'مسیر کامل'])

                # نوشتن داده‌ها
                for file_path_item in self.filtered_matches:
                    info = self._get_file_info(file_path_item)
                    writer.writerow([
                        info['name'],
                        info['type'],
                        info['size_str'],
                        info['modified_str'],
                        file_path_item
                    ])

            self._log_to_parent(f"💾 فایل CSV ذخیره شد: {file_path}", "success")
            QMessageBox.information(
                self,
                "موفق",
                f"لیست {len(self.filtered_matches)} فایل با موفقیت ذخیره شد."
            )

        except Exception as e:
            self._log_to_parent(f"❌ خطا در ذخیره CSV: {e}", "error")
            QMessageBox.critical(self, "خطا", f"خطا در ذخیره فایل:\n{e}")

    def export_to_excel(self):
        """Export لیست فایل‌ها به فرمت Excel"""
        if not self.filtered_matches:
            QMessageBox.information(self, "اطلاع", "فایلی برای Export وجود ندارد.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            QMessageBox.critical(
                self,
                "خطا",
                "کتابخانه openpyxl نصب نیست.\n"
                "لطفاً با دستور زیر آن را نصب کنید:\n\n"
                "pip install openpyxl"
            )
            return

        # انتخاب مسیر ذخیره
        default_name = f"ISO_Files_{self.line_no.replace('/', '-')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "ذخیره به Excel",
            default_name,
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ISO Files"

            # استایل هدر
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # نوشتن هدر
            headers = ['ردیف', 'نام فایل', 'نوع', 'حجم', 'تاریخ تغییر', 'مسیر کامل']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # نوشتن داده‌ها
            for row_idx, file_path_item in enumerate(self.filtered_matches, 2):
                info = self._get_file_info(file_path_item)

                ws.cell(row=row_idx, column=1, value=row_idx - 1)
                ws.cell(row=row_idx, column=2, value=info['name'])
                ws.cell(row=row_idx, column=3, value=info['type'])
                ws.cell(row=row_idx, column=4, value=info['size_str'])
                ws.cell(row=row_idx, column=5, value=info['modified_str'])
                ws.cell(row=row_idx, column=6, value=file_path_item)

            # تنظیم عرض ستون‌ها
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 80

            # ذخیره فایل
            wb.save(file_path)

            self._log_to_parent(f"📊 فایل Excel ذخیره شد: {file_path}", "success")
            QMessageBox.information(
                self,
                "موفق",
                f"لیست {len(self.filtered_matches)} فایل با موفقیت ذخیره شد."
            )

        except Exception as e:
            self._log_to_parent(f"❌ خطا در ذخیره Excel: {e}", "error")
            QMessageBox.critical(self, "خطا", f"خطا در ذخیره فایل:\n{e}")

    # ===== Context Menu =====

    def _show_context_menu(self, position):
        """نمایش منوی راست‌کلیک"""
        if self.table.rowCount() == 0:
            return

        menu = QMenu(self)

        # اکشن‌های اصلی
        open_action = menu.addAction("📂 باز کردن")
        open_action.triggered.connect(self.open_selected_files)

        open_folder_action = menu.addAction("📁 باز کردن پوشه")
        open_folder_action.triggered.connect(self.open_containing_folder)

        copy_action = menu.addAction("📋 کپی مسیر")
        copy_action.triggered.connect(self.copy_selected_paths)

        menu.addSeparator()

        # انتخاب
        select_all_action = menu.addAction("☑️ انتخاب همه")
        select_all_action.triggered.connect(self.table.selectAll)

        deselect_action = menu.addAction("⬜ حذف انتخاب")
        deselect_action.triggered.connect(self.table.clearSelection)

        menu.addSeparator()

        # Export
        export_csv_action = menu.addAction("💾 Export به CSV")
        export_csv_action.triggered.connect(self.export_to_csv)

        export_excel_action = menu.addAction("📊 Export به Excel")
        export_excel_action.triggered.connect(self.export_to_excel)

        # نمایش منو
        menu.exec(self.table.viewport().mapToGlobal(position))

    # ===== رویدادها =====

    def _on_double_click(self):
        """رویداد دابل کلیک روی سطر"""
        self.open_selected_files()

    def _on_selection_changed(self):
        """رویداد تغییر انتخاب"""
        selected_count = len(set(item.row() for item in self.table.selectedItems()))

        # فعال/غیرفعال کردن دکمه‌ها
        has_selection = selected_count > 0
        self.open_btn.setEnabled(has_selection)
        self.open_folder_btn.setEnabled(has_selection)
        self.copy_path_btn.setEnabled(has_selection)

        # بروزرسانی لیبل
        if selected_count == 0:
            self.selection_label.setText("هیچ فایلی انتخاب نشده")
        elif selected_count == 1:
            self.selection_label.setText("1 فایل انتخاب شده")
        else:
            self.selection_label.setText(f"{selected_count} فایل انتخاب شده")

    # ===== کمکی =====

    def _log_to_parent(self, message: str, level: str = "info"):
        """ارسال لاگ به پنجره والد"""
        if self.parent_window and hasattr(self.parent_window, 'log_to_console'):
            self.parent_window.log_to_console(message, level)
